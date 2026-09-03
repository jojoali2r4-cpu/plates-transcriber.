import os
import re
from google import genai
from groq import Groq
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import streamlit as st

# ضبط واجهة التطبيق
st.set_page_config(
    page_title="تفريغ لوحات السيارات", page_icon="🚗", layout="centered"
)


# دالة ذكية تستخدم Google Gemini أو Groq لتحليل النص وتنسيق اللوحات بدقة
def parse_text_with_ai(raw_text, api_key, provider="Gemini"):
    prompt = (
        "أنت مساعد ذكاء اصطناعي متخصص في تفريغ وتنظيم لوحات السيارات السودانية بدقة تامة.\n"
        "النص التالي هو تفريغ صوتي لـ Whisper يحتوي على مواقع، أرقام لوحات، وحروف (مثل أ، ب، م، ن، ر) وملاحظات (ن نقل، ت تاكسي).\n"
        "مهمتك:\n"
        "1. فك تداخل الكلمات واستخراج الحروف والأرقام لكل لوحة سيارة بشكل صحيح ومنفصل (مثال: إذا كان النص يذكر أرقام وحروف، حولها إلى شكل لوحة مثل: أ ب 4567 أو ب 1234).\n"
        "2. استخراج رقم الموقع الصحيح.\n"
        "3. استخراج التصنيف والملاحظات (ن، ت، ب، م، ف، ر).\n"
        "أعطني النتيجة حصراً في أسطر مفصولة بالرمز | بالترتيب التالي لكل سيارة:\n"
        "رقم الموقع | حروف وأرقام اللوحة | التصنيف والملاحظات\n"
        "لا تضف أي شروحات أو مقدمات، فقط الأسطر المطلوبة.\n\n"
        f"النص المراد تحليله:\n{raw_text}"
    )

    # المحاولة الأولى باستخدام Google Gemini (مستقر جداً وسريع ومجاني)
    if provider == "Gemini" or True:
        try:
            client = genai.Client(api_key=api_key)
            response = client.models.generate_content(
                model="gemini-2.5-flash", contents=prompt
            )
            if response and response.text:
                return response.text
        except Exception:
            pass

    # المحاولة البديلة باستخدام Groq في حال تطلب الأمر
    try:
        client = Groq(api_key=api_key)
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
        )
        if response.choices[0].message.content:
            return response.choices[0].message.content
    except Exception:
        pass

    return ""


# تحويل مخرجات الذكاء الاصطناعي إلى جدول بيانات مرتب
def create_dataframe_from_ai(ai_output):
    rows = []
    lines = ai_output.strip().split("\n")
    current_site = ""

    for line in lines:
        if "|" in line:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 3:
                site_raw = parts[0]
                plate = parts[1]
                classification = parts[2]

                site_num = "".join(re.findall(r"\d+", site_raw))
                if site_num:
                    current_site = site_num
                    site_val = current_site
                else:
                    site_val = ""

                if plate and plate != "-":
                    rows.append({
                        "plate": plate,
                        "site": site_val,
                        "classification": (
                            classification if classification != "-" else ""
                        ),
                    })

    # إظهار رقم الموقع لأول سيارة فقط في كل مجموعة
    seen_sites = set()
    final_rows = []
    for r in rows:
        s = r["site"]
        if s:
            if s not in seen_sites:
                seen_sites.add(s)
                final_rows.append(r)
            else:
                r_copy = r.copy()
                r_copy["site"] = ""
                final_rows.append(r_copy)
        else:
            final_rows.append(r)

    return pd.DataFrame(final_rows)


# بناء ملف Excel المنسق (يمين لليسار، 3 أعمدة)
def generate_excel(df, output_filename="تفريغ_اللوحات.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تفريغ اللوحات"
    ws.views.sheetView[0].rightToLeft = True

    headers = ["حروف وأرقام السيارة", "رقم الموقع", "التصنيف والملاحظات"]
    ws.append(headers)

    for _, row in df.iterrows():
        ws.append([row["plate"], row["site"], row["classification"]])

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    data_font = Font(name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_num in range(1, 4):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_num in range(2, len(df) + 2):
        for col_num in range(1, 4):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if cell.value == "" or cell.value is None:
                cell.value = None

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22

    wb.save(output_filename)
    return output_filename


# --- الواجهة التفاعلية للمستخدم ---
st.title("🚗 تطبيق تفريغ لوحات السيارات")
st.write(
    "ارفعي ملف التسجيل الصوتي للحصول على جدول Excel منظم ودقيق للوحات السيارات."
)

api_key = st.text_input(
    "أدخلي مفتاح API الخاص بك (Gemini API أو Groq API):", type="password"
)

uploaded_file = st.file_uploader(
    "اختاري ملف الصوت أو الريكورد من الجوال:", type=None
)

if uploaded_file is not None and api_key:
    with open("temp_audio_file.m4a", "wb") as f:
        f.write(uploaded_file.getbuffer())

    if st.button("بدء التفريغ والاستخراج"):
        # الخطوة 1: تفريغ الصوت عبر Whisper (باستخدام Groq لأجل الصوت)
        # ملاحظة: إذا كان مفتاحك من Google Gemini، سنحتاج مفتاح Groq للصوت فقط، أو يمكنك إدخال مفتاح Groq هنا.
        # لتسهيل الأمر، جعلناها تعتمد على مفتاح Groq في تفريغ الصوت ومفتاح Gemini للتحليل، أو العكس.
        # دعنا نطلب مفتاح Groq للصوت ومفتاح Gemini للتحليل إذا أردتِ، أو وضع مفتاحين.
        pass

    # للتسهيل المطلق ودون تعقيد المفاتيح، سنعتمد على نموذج Whisper المجاني لتفريغ الصوت عبر Groq ومفتاح واحد:
    groq_audio_key = st.text_input(
        "أدخلي مفتاح Groq API (لتفريغ الريكورد الصوتي):", type="password"
    )

    if uploaded_file is not None and groq_audio_key and api_key:
        if st.button("تشغيل المعالجة الشاملة"):
            with st.spinner("🎤 جاري تفريغ الصوت عبر Whisper وتحليله بالذكاء..."):
                try:
                    client_groq = Groq(api_key=groq_audio_key)
                    with open("temp_audio_file.m4a", "rb") as file:
                        transcription = client_groq.audio.transcriptions.create(
                            file=("audio.m4a", file.read()),
                            model="whisper-large-v3",
                            language="ar",
                            response_format="text",
                        )
                    raw_text = transcription
                except Exception as e:
                    st.error(f"خطأ في تفريغ الصوت: {e}")
                    raw_text = ""

            if raw_text:
                with st.spinner(
                    "📊 جاري استخراج وترتيب اللوحات في جدول Excel..."
                ):
                    ai_output = parse_text_with_ai(raw_text, api_key)
                    df = create_dataframe_from_ai(ai_output)
                    excel_file = generate_excel(df, "تفريغ_اللوحات.xlsx")

                st.success("✅ تمت المعالجة بنجاح! حملي الملف من الزر بالأسفل:")
                st.dataframe(df)

                with open(excel_file, "rb") as f:
                    st.download_button(
                        label="📥 تحميل ملف Excel الجاهز",
                        data=f,
                        file_name="تفريغ_اللوحات.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            else:
                st.error("تعذر إتمام التفريغ الصوتي.")
